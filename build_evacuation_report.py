"""
build_evacuation_pipeline_v2.py

Download current Oregon evacuation polygons, compare them with a previously
geocoded passenger-address CSV, and create a filtered, prioritized Excel call list.

The final list includes every address inside a zone plus addresses within a
configurable proximity threshold (10 miles by default).

Install:
    python -m pip install requests shapely pyproj openpyxl

Run:
    python build_evacuation_report.py data/passenger_address_coordinates.csv

Custom output:
    python build_evacuation_report.py data/passenger_address_coordinates.csv --output evacuation_call_report.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import importlib.util
import math
import os
import sys
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pyproj import Transformer
from shapely.geometry import Point, shape
from shapely.geometry.base import BaseGeometry
from shapely.ops import transform


LOGGER = logging.getLogger(__name__)
PIPELINE_VERSION = "2.1"

ARCGIS_ITEM_ID = "dd2bee51e3004dc4b1004df34fbd9e29"
ARCGIS_ITEM_URL = (
    "https://www.arcgis.com/sharing/rest/content/items/"
    f"{ARCGIS_ITEM_ID}"
)
FALLBACK_LAYER_URL = (
    "https://services.arcgis.com/uUvqNMGPm7axC2dD/arcgis/rest/services/"
    "Fire_Evacuation_Areas_Public/FeatureServer/0"
)

DEFAULT_OUTPUT_PATH = Path("data/evacuation_call_report.xlsx")
DEFAULT_ADDRESS_REPORT_PATH = Path("data/passenger_evacuation_report.csv")
DEFAULT_PASSENGER_SOURCE_PATH = Path("data/spedsta_passengers.csv")
DEFAULT_GEOCODER_MODULE_PATH = Path("data/geocode_passengers.py")
DEFAULT_GEOJSON_CACHE_PATH = Path("data/oregon_evacuation_areas.geojson")
DEFAULT_TIMEOUT_SECONDS = 60
DEFAULT_BATCH_SIZE = 200
DEFAULT_MAX_PROXIMITY_MILES = 10.0
METERS_PER_MILE = 1609.344
DISTANCE_CRS = "EPSG:5070"
WGS84_CRS = "EPSG:4326"

REQUIRED_ADDRESS_COLUMNS = {
    "passenger_name",
    "phone_numbers",
    "input_address",
    "longitude",
    "latitude",
}


class EvacuationReportError(RuntimeError):
    """Raised when live zone data or report processing fails."""


@dataclass(frozen=True)
class EvacuationZone:
    level: int
    geometry_wgs84: BaseGeometry
    geometry_distance: BaseGeometry
    incident_name: str = ""
    area_name: str = ""
    county: str = ""


@dataclass
class ReportRow:
    passenger_name: str
    phone_numbers: str
    address: str
    evacuation_zone: int | None
    proximity_distance_miles: float | None
    proximity_zone: int | None


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _request_json(
    session: requests.Session,
    url: str,
    *,
    method: str = "GET",
    params: dict[str, Any] | None = None,
    data: dict[str, Any] | None = None,
    timeout: int = DEFAULT_TIMEOUT_SECONDS,
) -> dict[str, Any]:
    try:
        if method.upper() == "POST":
            response = session.post(url, params=params, data=data, timeout=timeout)
        else:
            response = session.get(url, params=params, timeout=timeout)
        response.raise_for_status()
    except requests.RequestException as exc:
        raise EvacuationReportError(
            f"{method.upper()} request failed for {url}: {exc}"
        ) from exc

    try:
        result = response.json()
    except requests.exceptions.JSONDecodeError as exc:
        raise EvacuationReportError(
            f"The service returned invalid JSON for {response.url}"
        ) from exc

    if not isinstance(result, dict):
        raise EvacuationReportError(f"Expected a JSON object from {response.url}")

    if "error" in result:
        error = result["error"] if isinstance(result["error"], dict) else {}
        message = _clean(error.get("message")) or "Unknown ArcGIS error"
        details = error.get("details") or []
        if isinstance(details, list) and details:
            message += ": " + "; ".join(map(str, details))
        raise EvacuationReportError(f"ArcGIS error: {message}")

    return result


def resolve_layer_url(
    session: requests.Session,
    *,
    timeout: int = DEFAULT_TIMEOUT_SECONDS,
) -> str:
    try:
        item = _request_json(
            session,
            ARCGIS_ITEM_URL,
            params={"f": "json"},
            timeout=timeout,
        )
        service_url = _clean(item.get("url")).rstrip("/")
        if service_url:
            if service_url.rsplit("/", 1)[-1].isdigit():
                return service_url
            return f"{service_url}/0"
    except EvacuationReportError as exc:
        LOGGER.warning(
            "Could not resolve item metadata; using fallback layer URL: %s",
            exc,
        )

    return FALLBACK_LAYER_URL


def _get_object_ids(
    session: requests.Session,
    layer_url: str,
    *,
    where: str,
    timeout: int,
) -> list[int]:
    result = _request_json(
        session,
        f"{layer_url}/query",
        method="POST",
        data={
            "where": where,
            "returnIdsOnly": "true",
            "f": "json",
        },
        timeout=timeout,
    )

    object_ids = result.get("objectIds")
    if object_ids is None:
        return []
    if not isinstance(object_ids, list):
        raise EvacuationReportError("ArcGIS returned invalid objectIds data.")

    return sorted(value for value in object_ids if isinstance(value, int))


def download_oregon_evacuation_geojson(
    *,
    where: str = "1=1",
    batch_size: int = DEFAULT_BATCH_SIZE,
    timeout: int = DEFAULT_TIMEOUT_SECONDS,
    cache_path: str | Path | None = DEFAULT_GEOJSON_CACHE_PATH,
) -> dict[str, Any]:
    if batch_size < 1:
        raise ValueError("batch_size must be at least 1")

    with requests.Session() as session:
        session.headers.update({"User-Agent": "EvacuationCallReport/2.0"})
        layer_url = resolve_layer_url(session, timeout=timeout)
        LOGGER.info("Using Oregon evacuation layer: %s", layer_url)

        object_ids = _get_object_ids(
            session,
            layer_url,
            where=where,
            timeout=timeout,
        )
        LOGGER.info("Found %d current evacuation polygons.", len(object_ids))

        features: list[dict[str, Any]] = []

        for start in range(0, len(object_ids), batch_size):
            batch = object_ids[start : start + batch_size]
            response = _request_json(
                session,
                f"{layer_url}/query",
                method="POST",
                data={
                    "objectIds": ",".join(map(str, batch)),
                    "outFields": (
                        "OBJECTID,Fire_Name,Fire_Evacuation_Level,County,"
                        "Evac_Area_Name,HazardType,last_edited_date"
                    ),
                    "returnGeometry": "true",
                    "outSR": "4326",
                    "f": "geojson",
                },
                timeout=timeout,
            )

            batch_features = response.get("features")
            if response.get("type") != "FeatureCollection" or not isinstance(
                batch_features, list
            ):
                raise EvacuationReportError(
                    "ArcGIS did not return a valid GeoJSON FeatureCollection."
                )
            features.extend(batch_features)

    geojson: dict[str, Any] = {
        "type": "FeatureCollection",
        "features": features,
    }

    if cache_path is not None:
        path = Path(cache_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = path.with_suffix(path.suffix + ".tmp")
        temp_path.write_text(
            json.dumps(geojson, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        temp_path.replace(path)

    return geojson


def load_zones(geojson: dict[str, Any]) -> list[EvacuationZone]:
    transformer = Transformer.from_crs(
        WGS84_CRS,
        DISTANCE_CRS,
        always_xy=True,
    )

    zones: list[EvacuationZone] = []

    for feature in geojson.get("features", []):
        if not isinstance(feature, dict):
            continue

        properties = feature.get("properties") or {}
        geometry_data = feature.get("geometry")
        if not geometry_data:
            continue

        try:
            level = int(properties.get("Fire_Evacuation_Level"))
        except (TypeError, ValueError):
            continue

        if level not in {1, 2, 3}:
            continue

        try:
            geometry_wgs84 = shape(geometry_data)
        except Exception:
            continue

        if geometry_wgs84.is_empty:
            continue

        if not geometry_wgs84.is_valid:
            geometry_wgs84 = geometry_wgs84.buffer(0)

        if geometry_wgs84.is_empty:
            continue

        geometry_distance = transform(transformer.transform, geometry_wgs84)

        zones.append(
            EvacuationZone(
                level=level,
                geometry_wgs84=geometry_wgs84,
                geometry_distance=geometry_distance,
                incident_name=_clean(properties.get("Fire_Name")),
                area_name=_clean(properties.get("Evac_Area_Name")),
                county=_clean(properties.get("County")),
            )
        )

    return zones



def ensure_address_report(
    address_csv: str | Path,
    *,
    passenger_source_csv: str | Path = DEFAULT_PASSENGER_SOURCE_PATH,
    geocoder_module_path: str | Path = DEFAULT_GEOCODER_MODULE_PATH,
    timeout: int = DEFAULT_TIMEOUT_SECONDS,
) -> Path:
    """Return an existing geocoded report, or generate it automatically."""
    address_path = Path(address_csv)
    if address_path.exists():
        return address_path

    source_path = Path(passenger_source_csv)
    module_path = Path(geocoder_module_path)

    if not source_path.exists():
        raise EvacuationReportError(
            f"Address report is missing: {address_path}. "
            f"Passenger source file is also missing: {source_path}."
        )

    if not module_path.exists():
        raise EvacuationReportError(
            f"Address report is missing: {address_path}. "
            f"Geocoding module is also missing: {module_path}."
        )

    LOGGER.info(
        "Address report %s was not found; generating it from %s.",
        address_path,
        source_path,
    )

    spec = importlib.util.spec_from_file_location(
        "spedsta_geocode_passengers",
        module_path,
    )
    if spec is None or spec.loader is None:
        raise EvacuationReportError(
            f"Could not load geocoding module: {module_path}"
        )

    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    try:
        spec.loader.exec_module(module)
    except Exception as exc:
        raise EvacuationReportError(
            f"Could not import geocoding module {module_path}: {exc}"
        ) from exc

    create_report = getattr(module, "create_passenger_geocode_report", None)
    if not callable(create_report):
        raise EvacuationReportError(
            f"Geocoding module {module_path} does not provide "
            "create_passenger_geocode_report()."
        )

    try:
        records = create_report(
            source_path,
            address_path,
            prepare_only=False,
            timeout=timeout,
        )
    except Exception as exc:
        raise EvacuationReportError(
            f"Could not generate {address_path}: {exc}"
        ) from exc

    if not address_path.exists():
        raise EvacuationReportError(
            f"Geocoding completed but did not create {address_path}."
        )

    matched = sum(
        1 for record in records
        if getattr(record, "geocode_status", "") == "MATCH"
    )
    LOGGER.info(
        "Generated %s with %d address records (%d matched).",
        address_path.resolve(),
        len(records),
        matched,
    )
    return address_path

def load_passenger_addresses(input_csv: str | Path) -> list[dict[str, str]]:
    path = Path(input_csv)

    try:
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            reader = csv.DictReader(file)
            columns = set(reader.fieldnames or [])
            missing = REQUIRED_ADDRESS_COLUMNS - columns

            if missing:
                raise EvacuationReportError(
                    "Address CSV is missing required columns: "
                    + ", ".join(sorted(missing))
                )

            return list(reader)

    except OSError as exc:
        raise EvacuationReportError(f"Could not read {path}: {exc}") from exc


def _parse_coordinate(value: str) -> float | None:
    text = _clean(value)
    if not text:
        return None

    try:
        number = float(text)
    except ValueError:
        return None

    if not math.isfinite(number):
        return None

    return number


def classify_address(
    longitude: float,
    latitude: float,
    zones: list[EvacuationZone],
    transformer: Transformer,
) -> tuple[int | None, float | None, int | None]:
    point_wgs84 = Point(longitude, latitude)
    point_distance = transform(transformer.transform, point_wgs84)

    containing_levels = [
        zone.level
        for zone in zones
        if zone.geometry_wgs84.covers(point_wgs84)
    ]

    if containing_levels:
        level = max(containing_levels)
        return level, 0.0, level

    if not zones:
        return None, None, None

    best_distance_meters: float | None = None
    best_level: int | None = None

    for zone in zones:
        distance_meters = point_distance.distance(zone.geometry_distance)

        if (
            best_distance_meters is None
            or distance_meters < best_distance_meters - 0.01
        ):
            best_distance_meters = distance_meters
            best_level = zone.level
        elif (
            abs(distance_meters - best_distance_meters) <= 0.01
            and (best_level is None or zone.level > best_level)
        ):
            best_level = zone.level

    if best_distance_meters is None:
        return None, None, None

    return None, best_distance_meters / METERS_PER_MILE, best_level


def build_report_rows(
    passenger_rows: Iterable[dict[str, str]],
    zones: list[EvacuationZone],
    *,
    max_proximity_miles: float = DEFAULT_MAX_PROXIMITY_MILES,
    passenger_source_csv: str | Path = DEFAULT_PASSENGER_SOURCE_PATH,
    geocoder_module_path: str | Path = DEFAULT_GEOCODER_MODULE_PATH,
) -> list[ReportRow]:
    """
    Include only actionable passenger addresses:

    1. Every located address inside an evacuation zone.
    2. Located addresses outside the zones but no more than
       max_proximity_miles from the nearest evacuation polygon.

    Records without coordinates and records farther away are omitted from the
    final call list. They belong in the separate manual-address review report.
    """

    if max_proximity_miles < 0:
        raise ValueError("max_proximity_miles cannot be negative")

    transformer = Transformer.from_crs(
        WGS84_CRS,
        DISTANCE_CRS,
        always_xy=True,
    )

    result: list[ReportRow] = []

    for source in passenger_rows:
        longitude = _parse_coordinate(source.get("longitude", ""))
        latitude = _parse_coordinate(source.get("latitude", ""))

        if longitude is None or latitude is None:
            continue

        current_zone, distance_miles, nearest_zone = classify_address(
            longitude,
            latitude,
            zones,
            transformer,
        )

        is_inside = current_zone in {1, 2, 3}
        is_nearby = (
            current_zone is None
            and distance_miles is not None
            and nearest_zone in {1, 2, 3}
            and distance_miles <= max_proximity_miles
        )

        if not (is_inside or is_nearby):
            continue

        result.append(
            ReportRow(
                passenger_name=_clean(source.get("passenger_name")),
                phone_numbers=_clean(source.get("phone_numbers")),
                address=_clean(source.get("input_address")),
                evacuation_zone=current_zone,
                proximity_distance_miles=distance_miles,
                proximity_zone=nearest_zone,
            )
        )

    def priority_key(row: ReportRow) -> tuple[Any, ...]:
        if row.evacuation_zone in {1, 2, 3}:
            # All contained addresses precede every proximity-only address.
            return (
                0,
                -row.evacuation_zone,
                0.0,
                row.passenger_name.casefold(),
                row.address.casefold(),
            )

        # For nearby addresses, prioritize the severity of the nearest zone,
        # then the shortest distance to that zone.
        return (
            1,
            -(row.proximity_zone or 0),
            row.proximity_distance_miles
            if row.proximity_distance_miles is not None
            else float("inf"),
            row.passenger_name.casefold(),
            row.address.casefold(),
        )

    result.sort(key=priority_key)
    return result


def _verify_xlsx(path: Path) -> None:
    if not path.exists() or path.stat().st_size < 1000:
        raise EvacuationReportError(
            f"Excel workbook was not created correctly: {path}"
        )

    if not zipfile.is_zipfile(path):
        raise EvacuationReportError(
            f"Output is not a valid XLSX ZIP package: {path}"
        )

    with zipfile.ZipFile(path, "r") as archive:
        corrupt_member = archive.testzip()
        if corrupt_member is not None:
            raise EvacuationReportError(
                f"Corrupt XLSX member detected: {corrupt_member}"
            )

    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        if "Evacuation Call List" not in workbook.sheetnames:
            raise EvacuationReportError(
                "Workbook is missing the Evacuation Call List sheet."
            )
        workbook.close()
    except EvacuationReportError:
        raise
    except Exception as exc:
        raise EvacuationReportError(
            f"Workbook failed the reopen test: {exc}"
        ) from exc


def write_excel_report(
    rows: list[ReportRow],
    output_xlsx: str | Path,
) -> Path:
    output_path = Path(output_xlsx)

    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evacuation Call List"
    sheet.sheet_view.showGridLines = False

    headers = [
        "Passenger Name",
        "Phone Number(s)",
        "Address",
        "Evacuation Zone",
        "Distance to Zone (mi)",
        "Nearest Zone",
    ]

    counts = {
        3: sum(1 for row in rows if row.evacuation_zone == 3),
        2: sum(1 for row in rows if row.evacuation_zone == 2),
        1: sum(1 for row in rows if row.evacuation_zone == 1),
        "near_3": sum(
            1
            for row in rows
            if row.evacuation_zone is None and row.proximity_zone == 3
        ),
        "near_2": sum(
            1
            for row in rows
            if row.evacuation_zone is None and row.proximity_zone == 2
        ),
        "near_1": sum(
            1
            for row in rows
            if row.evacuation_zone is None and row.proximity_zone == 1
        ),
    }

    sheet.merge_cells("A1:F1")
    sheet["A1"] = "Passenger Evacuation Call List"
    sheet["A1"].fill = PatternFill("solid", fgColor="334155")
    sheet["A1"].font = Font(
        bold=True,
        color="FFFFFF",
        size=16,
    )
    sheet["A1"].alignment = Alignment(
        horizontal="left",
        vertical="center",
    )
    sheet.row_dimensions[1].height = 30

    sheet.merge_cells("A2:F2")
    sheet["A2"] = (
        "Includes every address inside an evacuation zone and addresses within "
        "10 miles. Priority: inside Level 3, 2, 1; then nearby Level 3, 2, 1, "
        "with the closest addresses first within each nearby group."
    )
    sheet["A2"].fill = PatternFill("solid", fgColor="E2E8F0")
    sheet["A2"].font = Font(italic=True, color="334155")
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 32

    summary_labels = [
        ("H1", "Inside Level 3", counts[3], "FCE8E8"),
        ("H2", "Inside Level 2", counts[2], "FCEEDB"),
        ("H3", "Inside Level 1", counts[1], "FBF6D8"),
        ("H4", "Near Level 3", counts["near_3"], "FDF3EE"),
        ("H5", "Near Level 2", counts["near_2"], "FDF7EA"),
        ("H6", "Near Level 1", counts["near_1"], "FAF9ED"),
    ]

    for cell_ref, label, value, fill in summary_labels:
        label_cell = sheet[cell_ref]
        value_cell = sheet.cell(
            row=label_cell.row,
            column=label_cell.column + 1,
            value=value,
        )
        label_cell.value = label
        label_cell.fill = PatternFill("solid", fgColor=fill)
        value_cell.fill = PatternFill("solid", fgColor=fill)
        label_cell.font = Font(bold=True, color="334155")
        value_cell.font = Font(bold=True, color="0F172A")
        value_cell.alignment = Alignment(horizontal="center")

    header_row = 5

    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=header)
        cell.fill = PatternFill("solid", fgColor="475569")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    sheet.row_dimensions[header_row].height = 30

    bottom_border = Border(
        bottom=Side(style="thin", color="CBD5E1")
    )

    for row_number, report_row in enumerate(rows, start=header_row + 1):
        values = [
            report_row.passenger_name,
            report_row.phone_numbers,
            report_row.address,
            report_row.evacuation_zone,
            report_row.proximity_distance_miles,
            report_row.proximity_zone,
        ]

        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=column in {2, 3},
            )
            cell.border = bottom_border

        sheet.cell(row=row_number, column=2).number_format = "@"
        sheet.cell(row=row_number, column=5).number_format = "0.00"

        if report_row.evacuation_zone == 3:
            fill = "FCE8E8"
        elif report_row.evacuation_zone == 2:
            fill = "FCEEDB"
        elif report_row.evacuation_zone == 1:
            fill = "FBF6D8"
        elif report_row.proximity_zone == 3:
            fill = "FDF3EE"
        elif report_row.proximity_zone == 2:
            fill = "FDF7EA"
        elif report_row.proximity_zone == 1:
            fill = "FAF9ED"
        else:
            fill = "FFFFFF"

        if fill != "FFFFFF":
            row_fill = PatternFill("solid", fgColor=fill)
            for column in range(1, 7):
                sheet.cell(row=row_number, column=column).fill = row_fill

    last_row = max(header_row, header_row + len(rows))

    sheet.auto_filter.ref = f"A{header_row}:F{last_row}"
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.print_title_rows = f"1:{header_row}"
    sheet.print_area = f"A1:I{last_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    widths = {
        "A": 24,
        "B": 27,
        "C": 48,
        "D": 18,
        "E": 22,
        "F": 16,
        "G": 3,
        "H": 22,
        "I": 10,
    }

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    temp_fd, temp_name = tempfile.mkstemp(
        prefix="evacuation_report_",
        suffix=".xlsx",
        dir=output_path.parent,
    )
    os.close(temp_fd)
    temp_path = Path(temp_name)

    try:
        workbook.save(temp_path)
        _verify_xlsx(temp_path)
        temp_path.replace(output_path)
        _verify_xlsx(output_path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise

    return output_path


def create_evacuation_call_report(
    address_csv: str | Path,
    output_xlsx: str | Path = DEFAULT_OUTPUT_PATH,
    *,
    zone_cache_path: str | Path | None = DEFAULT_GEOJSON_CACHE_PATH,
    where: str = "1=1",
    batch_size: int = DEFAULT_BATCH_SIZE,
    timeout: int = DEFAULT_TIMEOUT_SECONDS,
    max_proximity_miles: float = DEFAULT_MAX_PROXIMITY_MILES,
    passenger_source_csv: str | Path = DEFAULT_PASSENGER_SOURCE_PATH,
    geocoder_module_path: str | Path = DEFAULT_GEOCODER_MODULE_PATH,
) -> list[ReportRow]:
    address_path = ensure_address_report(
        address_csv,
        passenger_source_csv=passenger_source_csv,
        geocoder_module_path=geocoder_module_path,
        timeout=timeout,
    )

    geojson = download_oregon_evacuation_geojson(
        where=where,
        batch_size=batch_size,
        timeout=timeout,
        cache_path=zone_cache_path,
    )

    zones = load_zones(geojson)
    passenger_rows = load_passenger_addresses(address_path)
    report_rows = build_report_rows(
        passenger_rows,
        zones,
        max_proximity_miles=max_proximity_miles,
    )
    write_excel_report(report_rows, output_xlsx)
    return report_rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a prioritized passenger evacuation Excel report."
    )
    parser.add_argument(
        "address_csv",
        type=Path,
        nargs="?",
        default=DEFAULT_ADDRESS_REPORT_PATH,
        help=(
            "Geocoded CSV produced by geocode_passengers.py. If missing, "
            "it is generated automatically. Default: "
            f"{DEFAULT_ADDRESS_REPORT_PATH}"
        ),
    )
    parser.add_argument(
        "--passenger-source",
        type=Path,
        default=DEFAULT_PASSENGER_SOURCE_PATH,
        help=(
            "Raw SPEDSTA passenger CSV used when the geocoded report is "
            f"missing. Default: {DEFAULT_PASSENGER_SOURCE_PATH}"
        ),
    )
    parser.add_argument(
        "--geocoder-module",
        type=Path,
        default=DEFAULT_GEOCODER_MODULE_PATH,
        help=(
            "Path to geocode_passengers.py. Default: "
            f"{DEFAULT_GEOCODER_MODULE_PATH}"
        ),
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help=f"Output Excel workbook. Default: {DEFAULT_OUTPUT_PATH}",
    )
    parser.add_argument(
        "--zone-cache",
        type=Path,
        default=DEFAULT_GEOJSON_CACHE_PATH,
        help="Where to save the downloaded live Oregon GeoJSON.",
    )
    parser.add_argument(
        "--where",
        default="1=1",
        help="Optional ArcGIS SQL filter.",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=DEFAULT_BATCH_SIZE,
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=DEFAULT_TIMEOUT_SECONDS,
    )
    parser.add_argument(
        "--max-proximity-miles",
        type=float,
        default=DEFAULT_MAX_PROXIMITY_MILES,
        help=(
            "Include located addresses outside evacuation zones only when "
            "they are within this many miles. Default: 10."
        ),
    )
    parser.add_argument(
        "--quiet",
        action="store_true",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s: %(message)s",
    )

    LOGGER.info("SPEDSTA evacuation pipeline version %s", PIPELINE_VERSION)

    try:
        rows = create_evacuation_call_report(
            args.address_csv,
            args.output,
            zone_cache_path=args.zone_cache,
            where=args.where,
            batch_size=args.batch_size,
            timeout=args.timeout,
            max_proximity_miles=args.max_proximity_miles,
            passenger_source_csv=args.passenger_source,
            geocoder_module_path=args.geocoder_module,
        )
    except (EvacuationReportError, ValueError) as exc:
        LOGGER.error("%s", exc)
        return 1

    output_path = args.output
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")

    print(f"Wrote {len(rows)} records to {output_path.resolve()}")
    print("Workbook passed ZIP-integrity and reopen validation.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())