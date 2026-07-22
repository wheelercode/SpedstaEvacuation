"""
build_manual_address_review_fixed.py

Create an Excel workbook of passenger addresses that require manual correction.

Install:
    python -m pip install openpyxl

Run:
    python build_manual_address_review_fixed.py data/passenger_evacuation_report.csv

Custom output:
    python build_manual_address_review_fixed.py data/passenger_evacuation_report.csv --output manual_address_review.xlsx
"""

from __future__ import annotations

import argparse
import csv
import os
import tempfile
import zipfile
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


REVIEW_STATUSES = {
    "REVIEW",
    "NO_MATCH",
    "INVALID_ADDRESS",
    "MAILING_ONLY",
    "NO_ADDRESS",
}

STATUS_ORDER = {
    "REVIEW": 0,
    "NO_MATCH": 1,
    "INVALID_ADDRESS": 2,
    "MAILING_ONLY": 3,
    "NO_ADDRESS": 4,
}

STATUS_FILLS = {
    "REVIEW": "FFF7ED",
    "NO_MATCH": "FEFCE8",
    "INVALID_ADDRESS": "FEE2E2",
    "MAILING_ONLY": "EFF6FF",
    "NO_ADDRESS": "F1F5F9",
}

STATUS_LABELS = {
    "REVIEW": "Needs parsing cleanup",
    "NO_MATCH": "Geocoder returned no match",
    "INVALID_ADDRESS": "Invalid address value",
    "MAILING_ONLY": "Mailing address only",
    "NO_ADDRESS": "No address supplied",
}

OUTPUT_HEADERS = [
    "Status",
    "Passenger Name",
    "Phone Number(s)",
    "Original Address",
    "Parsed Street",
    "Parsed City",
    "State",
    "ZIP",
    "Address Source",
    "Reason / Notes",
    "Corrected Physical Address",
    "Correction Status",
]


class ManualReviewReportError(RuntimeError):
    """Raised when the manual-review workbook cannot be built."""


def load_review_rows(input_csv: str | Path) -> list[dict[str, str]]:
    input_path = Path(input_csv)

    try:
        with input_path.open("r", encoding="utf-8-sig", newline="") as file:
            reader = csv.DictReader(file)
            fieldnames = set(reader.fieldnames or [])

            required = {
                "geocode_status",
                "passenger_name",
                "phone_numbers",
                "input_address",
                "street",
                "city",
                "state",
                "zip",
                "address_source",
                "review_notes",
            }

            missing = required - fieldnames
            if missing:
                raise ManualReviewReportError(
                    "Input CSV is missing required columns: "
                    + ", ".join(sorted(missing))
                )

            rows = [
                dict(row)
                for row in reader
                if (row.get("geocode_status") or "").strip().upper()
                in REVIEW_STATUSES
            ]

    except OSError as exc:
        raise ManualReviewReportError(
            f"Could not read {input_path}: {exc}"
        ) from exc

    rows.sort(
        key=lambda row: (
            STATUS_ORDER.get(
                (row.get("geocode_status") or "").strip().upper(),
                99,
            ),
            (row.get("passenger_name") or "").casefold(),
            (row.get("input_address") or "").casefold(),
        )
    )

    return rows


def _verify_xlsx(path: Path) -> None:
    if not path.exists() or path.stat().st_size < 1000:
        raise ManualReviewReportError(
            f"Excel workbook was not created correctly: {path}"
        )

    if not zipfile.is_zipfile(path):
        raise ManualReviewReportError(
            f"Output is not a valid XLSX ZIP package: {path}"
        )

    with zipfile.ZipFile(path, "r") as archive:
        corrupt_member = archive.testzip()
        if corrupt_member is not None:
            raise ManualReviewReportError(
                f"Corrupt XLSX member detected: {corrupt_member}"
            )

    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        required_sheets = {"Manual Address Review", "Instructions"}

        if not required_sheets.issubset(set(workbook.sheetnames)):
            raise ManualReviewReportError(
                "Workbook is missing one or more required sheets."
            )

        workbook.close()

    except ManualReviewReportError:
        raise
    except Exception as exc:
        raise ManualReviewReportError(
            f"Workbook failed the reopen test: {exc}"
        ) from exc


def build_workbook(rows: list[dict[str, str]]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Manual Address Review"
    sheet.sheet_view.showGridLines = False

    sheet.merge_cells("A1:L1")
    sheet["A1"] = "Passenger Addresses Requiring Manual Review"
    sheet["A1"].fill = PatternFill("solid", fgColor="334155")
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=16)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30

    sheet.merge_cells("A2:L2")
    sheet["A2"] = (
        "These records do not have usable GIS coordinates. Enter a corrected "
        "physical address, update Correction Status, then update the source "
        "passenger file and rerun geocoding."
    )
    sheet["A2"].fill = PatternFill("solid", fgColor="E2E8F0")
    sheet["A2"].font = Font(italic=True, color="334155")
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 38

    counts = Counter(
        (row.get("geocode_status") or "").strip().upper()
        for row in rows
    )

    summary_headers = [
        "REVIEW",
        "NO MATCH",
        "INVALID",
        "MAILING ONLY",
        "NO ADDRESS",
        "TOTAL",
    ]
    summary_values = [
        counts.get("REVIEW", 0),
        counts.get("NO_MATCH", 0),
        counts.get("INVALID_ADDRESS", 0),
        counts.get("MAILING_ONLY", 0),
        counts.get("NO_ADDRESS", 0),
        len(rows),
    ]

    for column, value in enumerate(summary_headers, start=1):
        cell = sheet.cell(row=4, column=column, value=value)
        cell.fill = PatternFill("solid", fgColor="CBD5E1")
        cell.font = Font(bold=True, color="334155")
        cell.alignment = Alignment(horizontal="center")

    for column, value in enumerate(summary_values, start=1):
        cell = sheet.cell(row=5, column=column, value=value)
        cell.fill = PatternFill("solid", fgColor="F8FAFC")
        cell.font = Font(bold=True, color="0F172A", size=13)
        cell.alignment = Alignment(horizontal="center")

    header_row = 7
    first_data_row = 8

    for column, value in enumerate(OUTPUT_HEADERS, start=1):
        cell = sheet.cell(row=header_row, column=column, value=value)
        cell.fill = PatternFill("solid", fgColor="475569")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    sheet.row_dimensions[header_row].height = 31

    bottom_border = Border(
        bottom=Side(style="thin", color="D1D5DB")
    )

    for row_number, source in enumerate(rows, start=first_data_row):
        status = (
            source.get("geocode_status") or ""
        ).strip().upper()

        reason = (source.get("review_notes") or "").strip()
        if not reason:
            reason = STATUS_LABELS.get(status, status)

        values = [
            status,
            source.get("passenger_name", ""),
            source.get("phone_numbers", ""),
            source.get("input_address", ""),
            source.get("street", ""),
            source.get("city", ""),
            source.get("state", ""),
            source.get("zip", ""),
            source.get("address_source", ""),
            reason,
            "",
            "Needs review",
        ]

        row_fill = PatternFill(
            "solid",
            fgColor=STATUS_FILLS.get(status, "FFFFFF"),
        )
        edit_fill = PatternFill("solid", fgColor="ECFDF5")

        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = bottom_border
            cell.fill = row_fill if column <= 10 else edit_fill

        sheet.cell(row=row_number, column=3).number_format = "@"
        sheet.cell(row=row_number, column=8).number_format = "@"

    last_row = max(header_row, header_row + len(rows))

    if rows:
        validation = DataValidation(
            type="list",
            formula1=(
                '"Needs review,Corrected,No physical address,'
                'Unable to resolve"'
            ),
            allow_blank=False,
        )
        validation.error = "Select a value from the list."
        validation.errorTitle = "Invalid correction status"
        validation.prompt = "Choose the review status."
        validation.promptTitle = "Correction status"
        validation.showErrorMessage = True
        validation.showInputMessage = True
        sheet.add_data_validation(validation)
        validation.add(f"L{first_data_row}:L{last_row}")

    sheet.auto_filter.ref = f"A{header_row}:L{last_row}"
    sheet.freeze_panes = "A8"
    sheet.print_title_rows = f"1:{header_row}"
    sheet.print_area = f"A1:L{last_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    widths = {
        "A": 18,
        "B": 23,
        "C": 25,
        "D": 42,
        "E": 32,
        "F": 18,
        "G": 9,
        "H": 11,
        "I": 18,
        "J": 38,
        "K": 43,
        "L": 21,
    }

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    instructions = workbook.create_sheet("Instructions")
    instructions.sheet_view.showGridLines = False
    instructions.merge_cells("A1:D1")
    instructions["A1"] = "How to Correct the Address List"
    instructions["A1"].fill = PatternFill("solid", fgColor="334155")
    instructions["A1"].font = Font(bold=True, color="FFFFFF", size=15)
    instructions.row_dimensions[1].height = 28

    instruction_rows = [
        ["Step", "Action", "What to enter", "Expected result"],
        [
            1,
            "Review the original address and parser result.",
            "Use a reliable source or call the passenger.",
            "Confirm the physical residence.",
        ],
        [
            2,
            "Enter a complete physical address.",
            "Use Corrected Physical Address.",
            "Example: 411 E Carpenter Lane, Sisters, OR 97759",
        ],
        [
            3,
            "Update Correction Status.",
            "Choose Corrected or another final disposition.",
            "The list can be filtered by completion.",
        ],
        [
            4,
            "Update the passenger source file.",
            "Replace or add address 1/address 2.",
            "The source data becomes authoritative.",
        ],
        [
            5,
            "Rerun geocoding and the evacuation report.",
            "Use the corrected passenger CSV.",
            "Coordinates and evacuation proximity are rebuilt.",
        ],
    ]

    for row_number, values in enumerate(instruction_rows, start=3):
        for column, value in enumerate(values, start=1):
            cell = instructions.cell(
                row=row_number,
                column=column,
                value=value,
            )
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if row_number == 3:
                cell.fill = PatternFill("solid", fgColor="475569")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

    instruction_widths = {
        "A": 9,
        "B": 34,
        "C": 43,
        "D": 39,
    }

    for column, width in instruction_widths.items():
        instructions.column_dimensions[column].width = width

    instructions.freeze_panes = "A4"

    return workbook


def create_manual_review_report(
    input_csv: str | Path,
    output_xlsx: str | Path,
) -> int:
    rows = load_review_rows(input_csv)
    output_path = Path(output_xlsx)

    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook(rows)

    temp_fd, temp_name = tempfile.mkstemp(
        prefix="manual_address_review_",
        suffix=".xlsx",
        dir=output_path.parent,
    )
    os.close(temp_fd)
    temp_path = Path(temp_name)

    try:
        workbook.save(temp_path)
        _verify_xlsx(temp_path)
        temp_path.replace(output_path)
        _verify_xlsx(output_path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise

    return len(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Create an Excel workbook of passenger addresses requiring "
            "manual correction."
        )
    )
    parser.add_argument(
        "input_csv",
        type=Path,
        help="Completed passenger geocoding report CSV.",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("manual_address_review.xlsx"),
        help="Output Excel workbook.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    try:
        count = create_manual_review_report(
            args.input_csv,
            args.output,
        )
    except ManualReviewReportError as exc:
        print(f"ERROR: {exc}")
        return 1

    output_path = args.output
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")

    print(f"Wrote {count} records to {output_path.resolve()}")
    print("Workbook passed ZIP-integrity and reopen validation.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())